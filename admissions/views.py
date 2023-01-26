from django.shortcuts import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from .models import Admission
from .utils import admission_applicant_preview


# Doing it this way through a link request this is technically a completely open
# endpoint. Should wrap this in some permission decorator at some point
def download_callsheet_workbook(request):
    wb = Workbook()
    ws = wb.active

    admission = Admission.get_active_admission()
    ws.title = f"Ringeliste KSG {admission.semester}"

    # Name and phone number are usually both large cells
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    ws.cell(1, 1).value = "Navn"
    ws.cell(1, 1).font = Font(bold=True)

    ws.cell(1, 2).value = "Telefonnummer"
    ws.cell(1, 2).font = Font(bold=True)

    ws.cell(1, 3).value = "Skal få tilbud om"
    ws.cell(1, 3).font = Font(bold=True)

    ws.cell(1, 4).value = "Prioritet"
    ws.cell(1, 4).font = Font(bold=True)

    ws.cell(1, 5).value = "Fikk beholde"
    ws.cell(1, 5).font = Font(bold=True)

    ws.cell(1, 6).value = "Takket ja?"
    ws.cell(1, 6).font = Font(bold=True)

    ws.cell(1, 7).value = "Tidspunkter utilgjengelig"
    ws.cell(1, 7).font = Font(bold=True)
    ws.cell(2, 7).value = "Ikke implementert enda"

    # We use the graphql list as a proxy data model
    parsed_applicants = admission_applicant_preview(admission)

    for index, parsed_applicant in enumerate(parsed_applicants):
        ws.cell(index + 2, 1).value = parsed_applicant.full_name
        ws.cell(index + 2, 2).value = parsed_applicant.phone
        ws.cell(
            index + 2, 3
        ).value = parsed_applicant.offered_internal_group_position_name
        ws.cell(index + 2, 4).value = parsed_applicant.applicant_priority
        ws.cell(index + 2, 5).value = parsed_applicant.will_be_admitted

    response = HttpResponse(content_type="application/ms-excel")
    response[
        "Content-Disposition"
    ] = f"attachment; filename=Ringeliste KSG - {admission.semester}.xlsx"
    wb.save(response)
    return response
